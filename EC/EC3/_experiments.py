# Usage

## CV
# There are two ways:
#     - One use a dataframe based on EC Porridge format. This gives the cycles, times, voltages, and currents.
#         Use:
#                 cv = CV()
#                 cv.clean_porridge(df) #Data from the refined EC Porridge. 
#                 cv.load_porridge_cycle()
#     - Two use a dataframe having just the voltages and currents
#         Use:
#                 cv = CV(data = df, srates = (optional), extract_sr = True/False)


import pandas as pd
import re
import numpy as np
from scipy.signal import find_peaks
from scipy import integrate
import seaborn as sns


class QR:
    """
    Class for quick routines. 
    
    """
    def __init__(self):
        return
    #_________Inputs_______
    def input_list(self, sep = ",", type = 'int'):
        """Comma separated input string is turned to list"""
        a = input("Enter comma separated values:")
        a = a.split(sep = sep)
        a = [eval(type)(i) for i in a]
                   
        return a
    
    
    #__________List related__________
    def diff(self, li1, li2):
        "Find elements not common between lists"
        li_dif = [i for i in li1 + li2 if i not in li1 or i not in li2]
        return li_dif
    
    
    def list_1D_to_dict(self, source_list = None, categories = None, indices = None, by_index = False):
        """
        Usage
            Convert a list like: ['Name', 'a', 'b', 'c', 'Address', 1,2,3, 'Age', 2,4,5]
            to a dictionarly like: {'Name': ['a', 'b', 'c'], 'Address': [1,2,3], 'Age': [2,4,5]}

            if by_index = False:
                Divide source_list by categories and turn to dictionary.

            if by_index = True:
                Use indices to divide source_list and turn to dictionary using categories as keys.
                For this indices and categories must be in same order. 

        Args:
            source_list = list to categorize.
            categories = categories to follow
            by_loc = Use categories as name or as indices.
            
        Example:
            l = ['cata', 'b', 'c', 'catd', 'e', 'f', 'g', 'cath', 'i', 'j', 'k', 'catl', 'm']
            a = [0,3,7,11]

            d = list_1D_to_dict(source_list = l, categories = ['A', 'D', 'H', 'L'], indices = [0,3,7,11], by_index = True)
            d
        """
        import itertools

        cat_it = iter(categories)

        out = {}
        key = None
        if by_index ==False:
            for item in source_list:
                if item in categories:
                    out[item] = []
                    key = item
                else:
                    out[key].append(item) 


        else:
            b = indices[1:] + indices[:1]
            b[-1] = len(source_list)

            tups = list(zip(indices,b))

            for tup in tups:
                out[next(cat_it)] = source_list[slice(*tup)][1:]
        return out
    
    #___________Data Frames____________
    def fractionalize(self, df, frac = (0,1)):
        """
        Fractionalize dataframe. Might work on numpy arrays too.
        """
        if frac != (0,1):
            length = df.shape[0]
            df = pd.DataFrame(np.split(df, (int(length*frac[0]), int(length*frac[1])), axis = 0)[1])        
            
        return df
    def difference_between(self, li1, li2):
        """
        Find elements in one list not in the other.
        """
        li_dif = [i for i in li1 + li2 if i not in li1 or i not in li2]
        return li_dif

    def sort_df_cols_by_numbers(self, df):
        from natsort import natsorted
        sorted_cols = natsorted(df.columns)
        df = df.reindex(sorted_cols, axis=1)
        
        return df

    def find_closest_indices_in_column(self, df, value, colname):
        df[colname] =pd.to_numeric(df[colname])
        exactmatch = df[df[colname] == value]
        if not exactmatch.empty:
            return exactmatch.index
        else:
            lowerneighbour_ind = df[df[colname] < value][colname].idxmax()
            upperneighbour_ind = df[df[colname] > value][colname].idxmin()
            return [lowerneighbour_ind, upperneighbour_ind] 
        
    def find_closest_indices_in_all_columns(self, df, value):
        """
        Search all columns of 'df' for the indices closest to 'value'.
        The outer function calls a find_match which works for series and applies it on all columns.
        """
        def find_match(series):
            """

            """
            exactmatch = series[series == value]
            if not exactmatch.empty:
                return exactmatch.index
            else:

                lowerneighbour_ind = series[series< value].idxmax()
                upperneighbour_ind = series[series > value].idxmin()
                return [lowerneighbour_ind, upperneighbour_ind] 

        df = df.apply(pd.to_numeric)

        return df.apply(find_match)        
    
    #____________________Colors/Plotting________________
    def get_n_colors_from_map(self, palette = 'gist_gray', num = 15):
        #from pylab import *
        from matplotlib import colors
        import matplotlib.pyplot as plt


        cmap = plt.get_cmap(palette, num)
        color_list = [colors.rgb2hex(cmap(i)[:3]) for i in range(cmap.N)]
        #color_list
        print(color_list)
        display(sns.color_palette(color_list))

        return color_list


    def group_hue_and_style_from_dict(self, ax, first_category = None, second_category = None, mapping_dic = None):
        """
        Maps two categories of matplotib axes (e.g. Marker style with hue) so that all elements of one category have the same hue.
        Args:
            second_category is in labels. 
                e.g use: 
                    handles, labels = ax.get_legend_handles_labels() 
                    to get labels. labels = ['Category 1 Name', 'cat_1', 'cat_1', 'Category 2 Name', 'cat_2', 'cat_2']

            mapping_dic:
                Maps 'Category 2' elements to 'Category 1' i.e: {'cat_2': cat_1}

        """
        handles, labels = ax.get_legend_handles_labels()
        index_category2 = labels.index(second_category) #Find index of the second category (name) in ax labels.
        index_category1 = labels.index(first_category)
 
        #___Connect group and color____
        color_dict = {label: handle.get_facecolor()
                      for handle, label in zip(handles[1:index_category2], labels[1:index_category2])}

        # loop through the items, assign color via the subscale of the item idem
        for handle, label in zip(handles[index_category2 + 1:], labels[index_category2 + 1:]):
            handle.set_color(color_dict[mapping_dic[label]])

        # create a legend only using the items
        ax.legend(handles[index_category2 + 1:], labels[index_category2 + 1:], title=second_category,
                  bbox_to_anchor=(0.99, 1.0), fontsize=12, frameon = False)
        return

    #__________________Latexing________________
    def latexify(self, s):
        """
        Takes a string e.g. "Ti32C33" and returns a latexified version.
        Currently works on subscripts only. 
        
        """

        import re
        nums = re.findall(r'\d+', s)
        pos = [[m.start(0), m.end(0)] for m in re.finditer(r'\d+', s)]
        numpos = list(zip(nums, pos))
        #print(list(numpos))

        for num, pos in numpos:
            #print(num, pos)
            string = f"$_\mathrm{{{num}}}$"
            s = s[:pos[0]] + string + s[pos[1]:]

            for ind, (n, [p_st, p_end]) in enumerate(numpos):
                if p_st > pos[1]:
                    numpos[ind][1][0] += len(string)-len(num)
                    numpos[ind][1][1] += len(string)-len(num)

                pass

        return s
    
    #_______________Statistical________________
    def find_rmse(self, actual, estimate,
                  axis = 0, **kwargs
                 ):
        """
        Find rmse of two dataframes or numpy arrays with identical columns.
        Args:
            scaler: sklearn's scaler object. Use when you were comparing datasets across a study and one material is to be used as a scaler. 
        """


        return ((actual - estimate) ** 2).mean(axis = axis) ** .5
        

    
        
    
    
class ECBase:
    def __init__(self, active_mass = 0.005, area = 1, volume = 1):
        self.active_mass = active_mass
        self.area = area
        self.volume = volume
        pass
   
    def integrate(self, y, x, over_rows = (None, None), interval = (None, None), frac = (0,1), return_val = 'total', axis = 0):
        #print("integrate either by sample rows or by interval")
        """
        Integrate y over x. 
        
        Args:
            Integrate over one of:
            1. over_rows: 
            2. interval: over the interval values given by x. 
            3. frac: over the fraction.
            
            return_val: 'total' = Return trapezoidal answer. 
                        'cumulative' = Return cumulative results.
            
            area: 'math' for mathematical
                  'polygon' for polygon area
                    
            

        Returns:
            return_val: 'total' = Return trapezoidal answer. 
                        'cumulative' = Return cumulative results.
        
        """
        
        
        if over_rows != (None, None):
            print("WARNNG. over_rows not implemented")
            pass
        if interval != (None, None):
            print("WARNNG. Interval not implemented")
            pass
        if frac != (0,1):
            y = QR().fractionalize(y, frac)
            x = QR().fractionalize(x, frac)
        
        
        ################### trapz and cumtrapz givie different results when the range is -a to a
        ### Integrate by both 
        ### https://docs.scipy.org/doc/scipy-0.18.1/reference/generated/scipy.integrate.cumtrapz.html
        cumtrapz = integrate.cumtrapz(y, x, initial=0, axis = axis)
        if return_val == 'total':
            return cumtrapz[-1]
        
        if return_val =='cumulative': 
            return cumtrapz
        
        
        
        return
   
    def peaks(self, data, lookup_in = (None, None), add_offset = False, frac = (0,1), index_as_peaks = False, **kwargs ):
        """
        Use None for no limit on index
        
        Usage: 
            obj.find_current_peaks(lookup_in = (0,350), add_offset = True,  width = 1, height = 0.01, rel_height = 1)
            
            #___find peaks___
            for m in mat_df['Object']:
                print(m.name)
                m.cv.find_current_peaks(lookup_in = (0,350), add_offset = True,  width = 1, height = 0.01, rel_height = 1)
        
        Args:
            lookup_in: tuple of indices in which to look for peaks.
            frac: tuple of fractions of data in which to look for peaks. Overrides lookup_in.
            add_offset: If the peaks are not being searched from the beginning of the dataframe then the returned values will  start from where the subset starts. add_offset will count from the beginning of the original dataframe. 
            index_as_peaks: Use dataframe index as peak position. If index at the beginning is 750 then start counting from 750
        
        Returns:
            A dataframe of peaks.
        """
        
        df = pd.DataFrame()
        df_peak_find = data.iloc[:,:]
        if lookup_in:
            offset = lookup_in[0]
            lookup_in = slice(lookup_in[0], lookup_in[1])
            df_peak_find = data.iloc[lookup_in,:]
            
        if frac != (0,1):
            #Select the percent of rows from here to there.
            
            length = data.shape[0]
            df_peak_find = pd.DataFrame(np.split(data, 
                                                 (int(length*frac[0]), int(length*frac[1])), 
                                                 axis = 0)[1]
                                       )
            #print("Shape: ", df_peak_find.shape)
            
            pass
            
        for i in range(len(df_peak_find.columns)):
            peaks, properties = find_peaks(df_peak_find.iloc[:, i], **kwargs)
            df_pr = pd.DataFrame(properties)
            
            if index_as_peaks: #Use dataframe index as peak position. If index at the beginning is 750 then start counting from 750
                df_pr['peaks'] = df_peak_find.iloc[peaks].index
            else: #Else use ordinal position as peaks.
                df_pr['peaks'] = peaks
                
            df_pr['ID'] = df_peak_find.columns[i]
            
            if add_offset:
                add_offset_to = ['left_bases', 'right_bases', 'peaks', 'left_ips', 'right_ips']
                df_pr[add_offset_to] += offset

            df = pd.concat([df, df_pr])
        return df
    
    def polygon_area(self, p):
        """
        Shoelace Algorithm.
        
        For polygons with no self-crossings or internal holes.  It will return twice the _signed_ area of your polygon.
        
        If the vertices of the polygon are specified in counter-clockwise order (i.e. by the right-hand rule), then the area will be positive.  
        Otherwise, the area will be negative, assuming the polygon has non-zero area to begin with.
        
        https://stackoverflow.com/questions/451426/how-do-i-calculate-the-area-of-a-2d-polygon
        
        Args: 
            List/array of data in x,y pairings.
        """
        def segments(p):
            return zip(p, p[1:] + [p[0]])
        
        return 0.5 * abs(sum(x0*y1 - x1*y0
                             for ((x0, y0), (x1, y1)) in segments(p)))

    

class CV(ECBase):
    
    def __init__(self, data = None, srates = None, extract_sr = False):
        """
        MAKE SURE: The data has current columns containing 'Im' and scanrate in mV/s.
        'Im5', 'Im10' etc. And voltages 'Vf' as in 'Vf3', 'Vf5' etc.

        Args:
            data: pandas data frame containing voltages and currents.     
            srates: A list of scan rates.     
            extract_sr: If true then then the srates will be extracted from data.

        Creates:
            cv_legend: For use in plotting
            currents: Just the current columns
        
        """
    
        super().__init__()
        self.data = data
        self.srates = srates
        
        
        if self.data is not None:
            self._extract_currents_voltages(data, extract_sr)
            
        
        #______Empty Initializations_______
        self.current_peaks = self.lin_resids = self.quad_resids = self.lin_cap = self.lin_diff =  pd.DataFrame()
        self. lin_total = self.quad_diff = self.quad_cap = self.quad_total = pd.DataFrame()
        

       
        pass
    
    #___________Loading/Cleaning Data_____________
    def clean_porridge(self, df):
        #___Clean first row: Round off scan rates___
        df.iloc[0] = pd.to_numeric(df.iloc[0], errors='coerce')
        ind = df.iloc[0].dropna()
        df.iloc[0, ind.index] = [int(round(a)) for a in df.iloc[0, ind.index].to_list()]

        df.iloc[0, :] = df.iloc[0, :].fillna(method = "ffill")
        z = zip(df.iloc[0], df.iloc[2])
        s = []
        for i in z: s.append(str(i[1]) + str(i[0]))
        df.columns = s
        df.rename(columns = {list(df)[1]:'Pt', list(df)[0]: 'cycle'}, inplace=True)
        df.iloc[:, 0] = df.iloc[:, 0].fillna(method = "ffill")
        df.dropna(axis=0, how='any', thresh=None, subset=None, inplace=True)

        filt = ~((df['Pt'] == 'Pt') | (df['Pt']=='#'))
        df = df[filt]
        df = df.reset_index(drop = True).drop('Pt', axis = 1)
        df['cycle'] = df['cycle'].str.replace("CURVE", "")
        df['cycle'] = pd.to_numeric(df['cycle'])
        
        cols = df.columns.values
        cols = [c.replace(".0", "") for c in cols]
        df.columns = cols
        
        self.porridge_data = df
        
    def load_porridge_cycle(self, cycle):
        
        filt = (self.porridge_data['cycle'] == cycle)
        data = self.porridge_data[filt]
        
        #_______Extract voltages, currents, cycles______
        self._extract_currents_voltages(data, extract_sr = True)
         
    def _extract_currents_voltages(self, data, extract_sr = False):
        if extract_sr:
            #self.srates = [re.findall("\d+", s) for s in data.columns[2:]]
            self.srates = [re.findall("\d+", s) for s in data.filter(like='Im')]
            self.srates = [x for xs in self.srates for x in xs]
            self.srates = list(dict.fromkeys(self.srates)) 

        self.srates = [int(s) for s in self.srates]
        self.cv_legend = [str(s) + " mV/s" for s in self.srates]

        self.currents = data.filter(like='Im')
        self.voltages = data.filter(like='Vf')
        self.times = data.filter(like='T')

        #___Voltages and indices for each scan rate___
        self.v_window = pd.DataFrame(self.voltages.max() - self.voltages.min()).T
        self.v_switch = pd.DataFrame(self.voltages.max()).T
        
        
        vol = self.voltages
        vmax = self.v_switch.loc[0]
        v_idx = [vol.index[vol.iloc[:, i]==vmax[i]].tolist()[0] for i in range(len(vmax))]
        self.v_switch_idx = pd.DataFrame([v_idx], columns =self.v_switch.columns )
        
        if len(self.srates) != self.currents.shape[1]:
            raise Exception(f"{self.srates} does not match current columns {self.currents.columns}. Remove duplicate scanrates and load again.")
        #______OLD CODE ->> Does not work on porridge
        #self.v_switch_idx =  pd.DataFrame(self.voltages.idxmax()).T
        
        return
   
    #____________Finding metrics______________
    
    def find_capacitiveness(self, frac = (0, 0.5), model = 'Quadratic', take_abs = True):
        """
        Args:
            frac: What fraction of the values to use. Default = (0, 0.5) meaning, select forward cycle only.
            model: 'Linear' or 'Quadratic' model.
        """
        
        if model == 'Quadratic':
            qc = self.find_capacity(self.quad_cap, frac)
            qd = self.find_capacity(self.quad_diff, frac)
        elif model == 'Linear':
            qc = self.find_capacity(self.lin_cap, frac)
            qd = self.find_capacity(self.lin_diff, frac)
            
        capacitiveness_cols = [r'%Capacitive@'+ str(c)+r'mV/s' for c in self.srates]

        if take_abs:
            capacitiveness = qc.abs()/(qc.abs() + qd.abs())
        else: 
            capacitiveness = qc/(qc + qd)
        capacitiveness.columns = capacitiveness_cols
        return capacitiveness
    
    def find_capacitiveness_peak_method(self, which = 'forward'):
        """
        Finds capacitiveness by taking the ratio with an ideal supercapacitor.
        
        # Integrate  till switching potential
        Only works on forward scan for now.
        
        Question: Would it be better to simply compare with the maxima???
        
        """
        def find_all_current_peaks(currents):
            """
            Find all current peaks: For voltammograms that have no peaks, simply take the maximum value.
            
            """
            peaks = (self.current_peaks[['peak_heights', 'SR']]).T
            peaks.columns = ['Im' + str(int(i)) for i in peaks.loc['SR']]
            peaks.reset_index(inplace = True, drop = True)

            no_peaks_sr = ['Im' + str(i) for i in self.no_current_peaks]
            no_peaks_val = pd.DataFrame(currents[no_peaks_sr].max()).T
            
            
            return pd.concat([peaks, no_peaks_val], axis = 1).drop(index = 1)
        
        num_voltages = len(self.voltages.columns)
        
        print("Under development: Use with care!!!")
 
        currents = self.currents.copy()
        voltages = self.voltages.copy()
        vwindow = self.v_window

        
        if which == 'forward':
            frac = (0,0.5)
        if which == 'reverse':
            frac = (0.5,0)
        
        try:
            all_current_peaks = find_all_current_peaks(currents)
            
            #_________Sort columns_______
            all_current_peaks = np.array(QR().sort_df_cols_by_numbers(all_current_peaks))
            vwindow = np.array(QR().sort_df_cols_by_numbers(vwindow))
                        
            #____________________________
            
            if num_voltages == 1:
                area = self.integrate(y = currents, x= voltages, frac = frac)
            else:
                area = self.integrate(y = currents,x= voltages, frac = frac)


            ideal_supercap = vwindow * all_current_peaks

            self.capacitiveness = (area/ideal_supercap)
            self.capacitiveness = pd.DataFrame(self.capacitiveness,columns = currents.columns)

        except:    
            if self.current_peaks.empty:
                print("Find current_peaks first")
            # elif area.shape[0] != ideal_supercap.shape[0]:
            #     print(f"# Scan Rates = {area.shape[0]}\n# Peaks = {ideal_supercap.shape[0]}")
            #     print("Adjust current_peaks so that there is only one peak per scan rate. Consider removing indices close to switching potential")
        
        pass
  
    def find_capacitiveness_max_current_method(self, frac = (0,0.5)):
        """
        Find the maximum of each current. 
        Use the product of max_current and vwindow as the area of an ideal_capacitor
        Take the ratio of the integral of the area to find capacitiveness. 
        
        """
        
        
        currents = self.currents.copy()
        voltages = self.voltages.copy()
        srates = self.srates
        
        area = self.integrate(y = currents,x= voltages, frac = frac)
        
        

        vwindow = np.array(self.v_window)
        currents_max = np.array(pd.DataFrame(currents.max()).T)
        ideal_supercap = vwindow * currents_max
        
        capness_cols = [r'Cness@'+ str(c)+r'mV/s' for c in srates]     
        
     
        self.capacitiveness = pd.DataFrame((area/ideal_supercap), columns = capness_cols)

        
        
        return
    
    def find_capacity(self, currents = pd.DataFrame(), frac = (0,1)):
        """
        Integrate I over dV and divide by srates to find capacity.
        
        Warning:
            If no current is passed then the experimental current is passed AND the result is stored as an attribute.
            Otherwise if say cv.quad_cap is passed then the capacity is returned but no attribute is generated. 
        
        Args:

        """
        if frac == (0,1): print("Warning: Code computes mathematical area (not polygon). Consider using frac(0,0.5)")
        
        if currents.empty: currents = self.currents
        
        voltages = self.voltages
        srates = self.srates
        

        #_____Integrate____
        capacity = self.integrate(y = currents, x = voltages, frac = frac)/(np.array(srates)*10**-3)
        capacity_cols = [r'Q@'+ str(c)+r'mV/s' for c in srates]
        capacity = pd.DataFrame([capacity], columns = capacity_cols) 
        
        #___Return___
        if currents.equals(self.currents): 
            self.capacity = capacity
        else: 
            return capacity
        
        return
          
    
    def find_power(self, currents = pd.DataFrame(), frac = (0,1)):
        """
        Find power. If no currents are provided then the code uses the experimental data. 
        """
        if currents.empty: currents = self.currents
        
        voltages = self.voltages
        srates = self.srates
        
        power = self.integrate(y = currents, x = voltages, frac = frac)
        power_cols = [r'P@'+ str(c)+r'mV/s' for c in srates]     
        
        power = pd.DataFrame([power], columns = power_cols) 

        self.power = power
        
        if currents.equals(self.currents): 
            self.power = power
        else: 
            return power
        
        return
    
    def find_energy(self, currents = pd.DataFrame(), frac = (0,1)):
        """
        By default this calculates the energy of the associated experimental currents. For any other set (e.g. capacitive etc.) it returns the computed value. 
        """
        if currents.empty: currents = self.currents

        voltages = self.voltages
        srates = self.srates
        
        #___Find cumulative power first___
        
        power_cum = self.integrate(y= currents, x = voltages, frac = frac, return_val = 'cumulative')
        #___Integrate power to find energy___
        voltages = QR().fractionalize(voltages, frac)        
        energy = self.integrate(y = power_cum, x = voltages, frac = (0,1))/(np.array(srates)*(10**-3))
        energy_cols = [r'E@'+ str(c)+r'mV/s' for c in srates]     
        
        energy = pd.DataFrame([energy], columns = energy_cols) 

        if currents.equals(self.currents): 
            self.energy = energy
        else: 
            return energy
        
        return
    
    def find_current_peaks(self, lookup_in = (None, None), add_offset = False, index_as_peaks = True, frac = (0,1), **kwargs):
        """
        kwargs are the arguments for scipy.find_peaks()
        These include: height, rel_height, prominences etc.
        
        Args: 
            lookup_in: Indices in which to look up data in. 
            
        """

        self.current_peaks = self.peaks(self.currents, add_offset = add_offset, lookup_in = lookup_in, index_as_peaks = index_as_peaks, frac = frac, **kwargs)
        self.current_peaks.rename(columns = {"ID": "SR"}, inplace = True)
        self.current_peaks['SR'] = self.current_peaks['SR'].str.extract('(\d+)').astype(int)
        

        self.no_current_peaks = QR().difference_between(self.srates, self.current_peaks['SR'].tolist())

        
        pass
    
    #___________Regression____________
    def find_lin_reg_param(self):
        #create ibysqrt
        #find lin_reg_params k1 and k2
        from scipy.stats import linregress

        self.find_ibysqrtv()

        def calc_slope(row):
            a = linregress(np.sqrt(self.srates), y=row.astype(float))
            return a.slope, a.intercept#, a.rvalue, a.pvalue, a.stderr


        res = self.currents_ibysqrtv.apply(calc_slope, axis = 1)
        columns = ("k1 k2").split() #Slope, intercept
        self.lin_reg_params = pd.DataFrame([[a, b] for a,b in res.values], columns=columns)
        pass
    
    def find_quad_reg_param(self):
        #create ibysqrt
        #find quad_reg_params k1 and k2

        self.find_ibysqrtv()
       
        x = np.sqrt(self.srates)
        y = np.transpose(np.array(self.currents_ibysqrtv)) #np.polyfit needs one dataset per COLUMN rather than row.
        res = np.polyfit(x.astype(float),y.astype(float), 2, rcond=None, full=False, w=None, cov=False)
        res = np.transpose(res)


        # res = self.currents_ibysqrtv.apply(calc_slope, axis = 1)
        columns = ("(v) (v^1/2) (v^0)").split() #b3 b1 b2

        self.quad_reg_params = pd.DataFrame(res, columns=columns)

        
        pass
        
    def find_lin_cap_diff(self):
        """
        Find capacitive and diffusive by the the linear (k1,k2) model.
        
        Creates:
        cap_lin
        diff_lin
        """
        self.find_lin_reg_param() #Find parameters
        
        k1 =(self.lin_reg_params['k1']) 
        k2 =(self.lin_reg_params['k2'])
        v = self.srates
        vsqrt = np.sqrt(self.srates)
        
        self.lin_cap = pd.DataFrame(np.outer(k1, v), columns =  self.currents.columns, index =self.currents.index)
        self.lin_diff = pd.DataFrame(np.outer(k2, vsqrt), columns =  self.currents.columns, index =self.currents.index)
        self.lin_total = pd.DataFrame(self.lin_cap + self.lin_diff, columns = self.currents.columns, index =self.currents.index)
        self.lin_resids =pd.DataFrame(self.currents - self.lin_total, columns = self.currents.columns, index =self.currents.index)
        
        pass
    
    def find_quad_cap_diff(self):
        """
        Find capacitive and diffusive by the quadratic (b1,b2, b3) model.
        b2 and b3 are powers 1/2 and 3/2 of the scan rate.
        b1 is power 1
        
        Creates:
        cap_quad
        diff_quad
        """
        self.find_quad_reg_param()
        
        b3, vsqrt_cube  =(self.quad_reg_params['(v)']), np.array(self.srates)**(3/2)
        b1, v =(self.quad_reg_params['(v^1/2)']), self.srates
        b2, vsqrt = (self.quad_reg_params['(v^0)']), np.sqrt(self.srates)

        
        self.quad_cap = pd.DataFrame(np.outer(b1, v), columns = self.currents.columns, index =self.currents.index)
        self.quad_diff =pd.DataFrame(np.outer(b2, vsqrt) + np.outer(b3,vsqrt_cube), columns =  self.currents.columns, index =self.currents.index) 
        self.quad_total = pd.DataFrame(self.quad_cap + self.quad_diff, columns = self.currents.columns, index =self.currents.index)
        self.quad_resids =pd.DataFrame(self.currents - self.quad_total, columns = self.currents.columns, index =self.currents.index)

        
        
        pass
    
    def find_ibysqrtv(self):
        self.currents_ibysqrtv = np.array(self.currents)/(np.sqrt(self.srates))
        self.currents_ibysqrtv = pd.DataFrame(self.currents_ibysqrtv, columns = self.currents.columns, index = self.currents.index)
        
        return 
    
    def find_residuals_at(self, which = 'lin', at_voltages = [0.2, 0.4, 0.6], sr = None):
        """
        Use to make a residual plot. 
        
        Args: 
        
        Creates:
            residuals_at: Gives voltages and currents with a column 'Scan' describing whether we're on the forward scan or reverse.
        
        """
        def df_resids_tidy(df, at_voltages = None):
            indices = residual = voltages = []

            df_resids_at_all_v = pd.DataFrame()
            for v in at_voltages:
                neighbors = QR().find_closest_indices_in_all_columns(df, v) #Find closest neighbors
                neighbors.columns = self.srates 
                neighbors = neighbors.transform(np.sort) #Sort all columns so that smaller index is first
                neighbors['Scan'] = pd.Series(['Forward', 'Reverse']) #Smaller index is Forward, larger is Reverse. 
                neighbors['Voltage'] = v #Add the voltage. 

                value_vars = QR().diff(list(neighbors.columns), ['Voltage']) 
                neighbors = neighbors.melt(id_vars = ['Voltage', 'Scan'], value_vars = value_vars, var_name = 'SR', value_name = 'index')
                neighbors['Current'] = 'Im'+neighbors['SR'].astype(str)

                tl = list(zip(neighbors['index'], neighbors['Current']))
                neighbors['Residual'] = [current_fit.loc[i[0], i[1]] for i in tl]

                neighbors.drop(['Current', 'index'], inplace = True, axis = 1)

                df_resids_at_all_v = pd.concat([df_resids_at_all_v, neighbors], axis = 0)            
            
            return df_resids_at_all_v
        
        #____Ensure that at_voltages are in the potential window___
        min_at, max_at = min(at_voltages), max(at_voltages)
        min_v, max_v = self.voltages.iloc[:,0].min(), self.voltages.iloc[:,0].max()
        
        
        if min_at < min_v or max_at > max_v:
            print("At_voltages is out of CV bounds. Enter new voltages.")
            QR().input_list(type = 'float')

        #____Find resids if they're not already found____
        if which == 'lin':
            if self.lin_resids.empty: self.find_lin_cap_diff()
            current_fit = self.lin_resids
        elif which == 'quad':
            if self.quad_resids.empty: self.find_quad_cap_diff()
            current_fit = self.quad_resids
        
        #____Create residual dataframe for faceting____
        df_resids = df_resids_tidy(df = self.voltages, at_voltages = at_voltages)
        if which == 'lin':
            df_resids.index.name = 'LinearResidualsAt'
        elif which == 'quad':
            df_resids.index.name = 'QuadraticResidualsAt'
        


        self.residuals_at = df_resids
        return
    
    def find_rmse(self, which = None):
        """
        Find rmse

        Args:
            which = 'linear', 'quadratic', 'both'
        """
        m = self
        
        rmse_cols_lin = [r'RMSE Linear@'+ str(c)+r'mV/s' for c in self.srates] 
        rmse_cols_quad = [r'RMSE Quadratic@'+ str(c)+r'mV/s' for c in self.srates] 
        
        
            
            
        
        if which == 'linear':
            
            #rmse = pd.DataFrame(((m.currents - m.lin_total) ** 2).mean() ** .5).T
            rmse = pd.DataFrame(QR().find_rmse(m.currents, m.lin_total)).T
            
            rmse.index = ['linear']
            rmse.columns = rmse_cols_lin
            
            

        if which == 'quadratic':
            quad_rmse = pd.DataFrame(QR().find_rmse(m.currents, m.quad_total)).T
            rmse.index = ['quadratic']
            rmse.columns = rmse_cols_quad

        
        if which == 'both':
            lin_rmse = pd.DataFrame(QR().find_rmse(m.currents, m.lin_total)).T
            quad_rmse = pd.DataFrame(QR().find_rmse(m.currents, m.quad_total)).T
            
            rmse = pd.concat([lin_rmse, quad_rmse], axis = 1)
            
            
            
            
            #rmse.index = ['linear', 'quadratic']
            rmse.columns = rmse_cols_lin+rmse_cols_quad

        

        self.rmse = rmse
        
        return 
    #___________Convenience/Shortcuts___________
    def drop_scanrates(self, drop = None):
        for rate in drop:
            self.currents.drop(list(self.currents.filter(regex = str(rate))), axis = 1, inplace = True)
            self.voltages.drop(list(self.voltages.filter(regex = str(rate))), axis = 1, inplace = True)
            self.times.drop(list(self.times.filter(regex = str(rate))), axis = 1, inplace = True)
            self.srates.remove(rate)
        
        return
    

    def set_scanrates(self, rates):
        """Set scan rates if necessary: Use with care because current columns must commensurate"""
        self.srates = rates
        return
    

class GCD(ECBase):
    def __init__(self, data = None, sp_currents = None, extract_sp_curr = False):
        """
        MAKE SURE: The data has current columns containing 'Im' and scanrate in mV/s.
        'Im5', 'Im10' etc. And voltages 'Vf' as in 'Vf3', 'Vf5' etc.

        Args:
            data: pandas data frame containing voltages and currents.     
            sp_currents: A list of specific currents.     
            extract_sr: If true then then the sp_currents will be extracted from data.

        Creates:
            cv_legend: For use in plotting
            currents: Just the current columns
        
        """
    
        super().__init__()
        self.data = data
        self.sp_currents = sp_currents
        
        #_______Extract currents and current densities______
        if self.data is not None:
            self._extract_currents_voltages(data, extract_sp_curr)


        
        

        
        pass
    
    def _extract_currents_voltages(data, extract_sp_curr):
        #self.raw = data
        
        if extract_sp_curr:
            self.sp_currents = [re.findall("\d+", s) for s in self.data.filter(like='Vf')]
            self.sp_currents = [x for xs in self.sp_currents for x in xs]
            self.sp_currents = list(dict.fromkeys(self.sp_currents)) 
        
        self.sp_currents = [float(s) for s in self.sp_currents]
        
        self.gcd_legend = [str(s) + " A/s" for s in self.sp_currents]
        
        self.voltages = self.data.filter(like='Vf')
        self.times = self.data.filter(like='Common Time') #Try to do it without common time. 
        
                
        #___Voltages and indices for each scan rate___
        self.v_window = pd.DataFrame(self.voltages.max() - self.voltages.min()).T

        
        self.t_switch_idx =  pd.DataFrame(self.voltages.idxmax()).T #Index at which discharge starts 
        self.t_switch = pd.DataFrame(self.times.to_numpy()[self.t_switch_idx, np.arange(len(self.times.columns))], columns = self.voltages.columns) #time at which discharge starts
        
        
        return
    
    
    
    def set_sp_currents(self, sp_currents):
        """Set scan rates if necessary: Use with care because current columns must commensurate"""
        self.sp_currents = sp_currents
    
    def find_discharge_voltages(self, offset = True):
        """
        Create discharging voltages. 
        """
        length = self.times.shape[0]
        ind = self.t_switch_idx.loc[0]
        
        #_______Find Discharge VOltages__________
        ml = max([ length - i for i in ind ]) #maximum length
        data = np.array(self.voltages)



        result = np.array([ np.pad(data[rInd:, i[0]], (0,  ml - length + rInd),constant_values = np.nan) for i, rInd in np.ndenumerate(ind) ]).T
        result = result[~np.isnan(result).all(axis=1)] #Remove rows that are all nan.
        
        self.discharge_voltages = pd.DataFrame(result, columns = self.voltages.columns)
        
        
        #_______FIND DISCHARGE TIMES TO ALIGN WITH DISCHARGE_VOLTAGES______
        
        d = self.discharge_voltages.shape[0]
        self.discharge_times = self.times.iloc[0:d,:]


        
        pass
    
    def find_power(self, which = 'specific'):
        which_presets = ['specific', 'areal', 'volumetric']
        pass
    def find_energy(self):
        pass
    def find_capacity(self):
        #Charging and discharging.
        pass
    def find_charge_discharge_time_ratio(self):
        pass
    
    def stretch_interpolate_uneven(self, data):
        #https://stackoverflow.com/questions/73741386/python-stretch-and-interpolate-time-series-of-unequal-length/73742290#73742290
        
        data = self.raw
        
        to = data.shape[0]

        tb_resized  = np.repeat(np.nan,to)
        foreign = np.linspace(0,to-1,len(tb)).round().astype(int)
        tb_resized[foreign] = tb

        bv_resized  = np.repeat(np.nan,to)
        foreign = np.linspace(0,to-1,len(tb)).round().astype(int)
        bv_resized[foreign] = bv
        d = {'ta': ta, 'a_val': av, 'tb':tb_resized, 'b_val':bv_resized}

        df_even = pd.DataFrame(dict([ (k,pd.Series(v)) for k,v in d.items()])).interpolate()

        return df_even
    
    def from_porridge_RAW(self, df_gcd):
        #_______Functions_________
        def clean_porridge_RAW(df_gcd):
            """
            Clean the GCDData(Raw) styled dataframe from porride. 

            Creates:


            """
            df_gcd_raw = df_gcd.copy()
            start_row = df_gcd.loc[df_gcd[1]=='CURVE'].index.tolist()
            end_row = df_gcd.loc[df_gcd[1]=='STARTTIMEOFFSET'].index.tolist()

            df_gcd_raw.iloc[:end_row[0]+1, 1] = 'Charging'
            df_gcd_raw.iloc[end_row[0]+1:, 1] = 'Discharging'

            df_gcd_raw = df_gcd_raw.iloc[start_row[0]+3:, :]
            df_gcd_raw.drop(df_gcd_raw.columns[0], inplace = True, axis = 1)
            df_gcd_raw.columns = ['cycle'] + df_gcd.iloc[start_row[0]+1, 2:].tolist() # Add 'cycle' to the list of columns

            #_____Create labels_____
            currents = df_gcd_raw[['Im']].iloc[0].astype(float).round(4)
            currents_rpt = currents.repeat(2)
            currents_rpt = pd.concat([pd.Series(['']),currents_rpt])

            #_____Select columns_____
            df_gcd_raw = df_gcd_raw.filter(regex='cycle|^Vf$|^T$')
            df_gcd_raw.columns = [(str(c) + str(curr)) for (c, curr) in zip(df_gcd_raw.columns, currents_rpt)];



            #____Separate df_ch and df_dis
            df_ch = df_gcd_raw[(df_gcd_raw.cycle == 'Charging')]
            df_ch.reset_index(drop = True, inplace = True)

            df_dis = df_gcd_raw[(df_gcd_raw.cycle == 'Discharging')]

            start_row = df_dis.loc[df_dis.iloc[:, 1]=='s'].index[0]
            df_dis = df_dis.loc[start_row + 1:,:]
            df_dis.reset_index(drop =True, inplace = True)




            def QR_remove_last_non_Nan(df, cols): #QR = Quick Routines
                rows = [i.loc[:, j].last_valid_index() for j in cols]
                z = zip(rows, cols)
                for e in z:
                    i.loc[e[0], e[1]] = np.NaN
                return i

            #_________Remove nuisance strings.
            for i in [df_ch, df_dis]:
                #_____i. Remove offset time.
                cols = []
                cols = i.filter(regex=("T.*")).columns.tolist()
                QR_remove_last_non_Nan(i, cols)

                #_____i. Remove certain strings.
                nuisances = ['STARTTIMEOFFSET', 'QUANT', r'Start Time Offset (s)']
                i.replace(nuisances, np.NaN, inplace = True)

            df_ch = df_ch[:-1] #Drop last row
            df_dis = df_dis[:-1]

            #Cleaning Done
            #_____________________________________



            return df_ch, df_dis, currents
        def find_cycle_times(df_ch, df_dis):
            def QR_return_last_non_Nan(df, cols): #QR = Quick Routines
                rows = [df.loc[:, j].last_valid_index() for j in cols]
                z = zip(rows, cols)

                last = []
                for e in z:
                    last.append(df.loc[e[0], e[1]])
                last = pd.DataFrame([last], columns = cols)
                return last

            cols = df_ch.filter(regex=("T.*")).columns.tolist()
            tch = QR_return_last_non_Nan(df_ch, cols )
            tdis = QR_return_last_non_Nan(df_dis, cols )


            cycle_times = pd.concat([tch, tdis], axis = 0)
            cycle_times.index = ['Charging', 'Discharging']

            #___Created___
            return cycle_times
        def normalize_by_time_and_update_df(df_ch, df_dis, cycle_times):
            df_ch.update(df_ch.filter(regex=("T.*"))/cycle_times.loc['Charging'])
            df_dis.update(df_dis.filter(regex=("T.*"))/cycle_times.loc['Discharging'])
            
            return df_ch, df_dis
        def stretch_interpolate(df_ch, df_dis):
            #____Stretch____
            def stretch(df):
                to = df.shape[0]

                stretched = pd.DataFrame()
                for tb in df.columns:
                    length = df[tb].dropna().shape[0]

                    tb_resized  = np.repeat(np.nan,to) #Create array equal to largest.
                    foreign = np.linspace(0, to-1, length).round().astype(int) #Create array ranging from 0 to total length but generate elements equal to shorter length.


                    tb_resized[foreign] = (df[tb].dropna()) #Fill the larger array. There will be gaps. 
                    df_2 = pd.DataFrame(tb_resized, columns = [tb])

                    stretched = pd.concat([stretched, df_2], axis = 1)


                return stretched
            
            df_ch_str = stretch(df_ch.drop(['cycle'], axis = 1)) #exclude the cycle column.
            df_dis_str = stretch(df_dis.drop(['cycle'], axis = 1)) #exclude the cycle column.
            
            #____Interpolate____                     
            df_ch = pd.DataFrame(dict([ (k,pd.Series(v)) for k,v in df_ch_str.items()])).interpolate()
            df_dis = pd.DataFrame(dict([ (k,pd.Series(v)) for k,v in df_dis_str.items()])).interpolate()
            
            return df_ch, df_dis

        #_______Algorithm_________
        self.charging, self.discharging, self.currents = clean_porridge_RAW(df_gcd)
        self.cycle_times = find_cycle_times(self.charging, self.discharging)
        self.charging, self.discharging = normalize_by_time_and_update_df(self.charging, self.discharging, self.cycle_times)
        self.charging_stretched, self.discharging_stretched = stretch_interpolate(self.charging, self.discharging)
        

        return

class EIS(ECBase):
    def __init__(self, data = None):
        super().__init__()
        self.eis = pd.DataFrame()
        
        if data:
            self.eis = pd.DataFrame(data)

                        
    def from_excel(self, file, freq = [10**5, 1]):
        """
        Pass file with 'Zreal', 'Zimag' and 'freq'. 
        Change the column names according to those in the excel file.
        
        If frequencies are not available the code assumes the range 10^5 (start) to 1 (stop). 
        
        
        Returns:
            Nothing
            
        Creates:
            self.eis
        
        """
        colnames = ['Zreal', 'Zimag', 'freq']
        
        from openpyxl import Workbook
        import openpyxl

        wb = openpyxl.load_workbook(file, read_only=True)
        ws = wb.active


        coord = {}
        for row in ws.iter_rows():
            for cell in row:
                if cell.value in colnames:
                    coord[cell.value] = (cell.row, cell.column)
                    coord[cell.value] = (cell.row, cell.column)
        maxRow = ws.max_row

        data = ws.values
        columns = next(data)[0:]

        self.eis = pd.DataFrame(data, columns=columns)
        
        if not 'freq' in self.eis.columns:
            self.eis['freq'] = np.geomspace(start=freq[0], stop=freq[1], num=self.eis.shape[0])
            
            
        
          
        pass
    
    def eis_fitting(self, circuit ='R0-p(R1,C1)-p(R2-Wo1,C2)', initial_guess =[.01, .01, 100, .01, .05, 100, 1], preset = 'Custom'):
        """
        Uses impedance.py to find the fit for the EIS object using the circuit given. 
        
        If 'Randles' is given, fits a randles circuit with cpe and one without cpe. 
        
        
        Args:
        preset: ['Custom', 'Randles', 'RandlesCPE']
        initial_guess:
            - For Randles could be: [.01, .005, .001, 200, .1]
            - For RandlesCPE could be: [.01, .005, .001, 200, .1, .9]
        
        Creates:
        - circuit_fit
        - randles_fit
        - randlesCPE_fit
        
        
        
        """
        from impedance import preprocessing
        from impedance.models.circuits import Randles, CustomCircuit


        frequencies = np.array(self.eis['freq'])
        Z = np.array(self.eis['Zreal'] + 1j * self.eis['Zimag'])

        # keep only the impedance data in the first quandrant
        #frequencies, Z = preprocessing.ignoreBelowX(frequencies, Z)

        f_pred = frequencies
        if preset == 'Custom':
            circuit = CustomCircuit(circuit, initial_guess=initial_guess)
            circuit.fit(frequencies, Z)
            prediction = circuit.predict(frequencies)
            circuit_fit_real, circuit_fit_imag =prediction.real, prediction.imag
            self.circuit_fit = pd.DataFrame(zip(frequencies, circuit_fit_real, circuit_fit_imag), columns = ['freq', 'Zreal', 'Zimag'])
        
        if preset == 'Randles':
            randles = Randles(initial_guess=initial_guess)
            randles.fit(frequencies, Z)
            randles_fit = randles.predict(f_pred)
            self.randles_fit = pd.DataFrame(zip(frequencies,randles_fit.real, randles_fit.imag), columns = ['freq', 'Zreal', 'Zimag'])
            
        if preset == 'RandlesCPE':
            randlesCPE = Randles(initial_guess=initial_guess, CPE=True)            
            randlesCPE.fit(frequencies, Z)
            randlesCPE_fit = randlesCPE.predict(f_pred)
            self.randlesCPE_fit = pd.DataFrame(zip(frequenciesrandlesCPE_fit.real, randlesCPE_fit.imag), columns = ['freq', 'Zreal', 'Zimag'])
        
        
        pass

class Material(CV, GCD):
    
    def __init__(self, name:str, experiments = {}):
        self.name = name
        self.nameLTX = QR().latexify(name)
        #self.cv = CV(data = None)
        if "CV" in experiments:
            self.cv = CV(data = None)
        else:
            pass
            #print("No data. Call clean_porridge followed by load_porridge_cycle if you have porridge data.")

        
    
        
            
#         if "GCD" in epxeriments:
#             pass
        
#         if "EIS" in experiments:
#             pass

        return

class Study(Material):
    """
    Study comprises Materials which comprise experiments
    
    A study may have Groups (Oxides, Sulfides for instance)
    """
    
    def __init__(self, data):
        """
        Data is a multindex dataframe which contains:
        
        Index: Group and Name i.e. Oxide and AgO
        Column: Material objects.
        
        """
        
        self.mats = data
        self.num_mats = data.shape[0]
        
        pass
    
    def extract_groups(self):
        
        self.groups = list(zip(*self.materials.index))[0]
        self.groups = list(set(self.groups))
    
    def load_porridge_folder(self):
        """
        Load porridge files from folder.
        """
        
        pass
    def Clean_CV_Porridge(self, cv_cycle = 1):
        """
        Untested
        
        CV: Sheet must be named CVData.
        study.mats must have file_paths
        
        """
        file_paths = self.mats['file_paths']
        
    
        for i in range(self.num_mats):
            m = self.mats.iloc[i,0]
            print(m.name)
            df = pd.read_excel(file_paths[i], engine = 'pyxlsb', sheet_name = 'CVData', header = None)
            m.cv.clean_porridge(df)
            #cycle = mat_df.iloc[i,0].cv.porridge_data['cycle'].max()
            m.cv.load_porridge_cycle(cv_cycle)
        return
    
    
    def study_subset(self, index = None, group_first = False, attribute = 'object'):
        """
        Subset study by index or by name.

        group_first overrides index.
        The study dataframe must have multi-indices. 

        Args:
            st = Study
            index = indices of materials to select
            group_first = True. Select the first element of each group as representative of that group.

            ret_type = 'object' or 'name'. If name then only the names are returned. 


            → MODIFY TO RETURN ANY ATTRIBUTE
        Returns:
            select: list of objects.
        """
        st = self
        select = []

        if attribute == 'object':
            if group_first == True:
                for j, (ind, group) in enumerate(st.mats.groupby(level=0)):
                    select.append(group.iloc[0,0])
            else:
                for i in index:
                    select.append(st.mats.iloc[i,0])

        else: 
            if group_first == True:
                for j, (ind, group) in enumerate(st.mats.groupby(level=0)):
                    select.append(getattr(group.iloc[0,0], attribute))
            else:
                for i in index:
                    select.append(getattr(st.mats.iloc[i,0], attribute))

        return select
    
class StudyCV(Study):
    """
    Collection of CV related study methods. 
    """
    def __init__():
        return
        
        
        
        
            